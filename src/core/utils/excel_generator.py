import io
from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy import desc
from db.models import SaleTransaction, Product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_sales_excel(db: Session, scenario_id: int) -> io.BytesIO:
    wb = Workbook()
    
    # -- Configurar estilos generales --
    font_family = "Segoe UI"
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=10, color="1F2937")
    font_total = Font(name=font_family, size=10, bold=True, color="1F2937")
    
    fill_header = PatternFill(start_color="D4A853", end_color="D4A853", fill_type="solid") # Oro corporativo
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")  # Gris ultra claro
    fill_total = PatternFill(start_color="FEFBF3", end_color="FEFBF3", fill_type="solid")  # Crema suave
    
    border_thin = Side(border_style="thin", color="E5E7EB")
    border_double = Side(border_style="double", color="374151")
    
    border_data = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_total = Border(top=border_thin, bottom=border_double)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # -- Hoja 1: Resumen de Cierres --
    ws1 = wb.active
    ws1.title = "Resumen de Cierres"
    ws1.views.sheetView[0].showGridLines = True
    
    # Cabeceras
    headers1 = ["Fecha", "Unidades Vendidas", "Ingresos Totales", "Dia de la Semana", "Es Quincena"]
    ws1.append(headers1)
    
    # Formatear fila de cabecera
    ws1.row_dimensions[1].height = 26
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Query para Resumen
    from sqlalchemy import func
    rows_summary = (
        db.query(
            SaleTransaction.sale_date,
            func.sum(SaleTransaction.quantity_sold).label("total_units"),
            func.sum(SaleTransaction.revenue).label("total_revenue"),
        )
        .filter(SaleTransaction.scenario_id == scenario_id)
        .group_by(SaleTransaction.sale_date)
        .order_by(desc(SaleTransaction.sale_date))
        .all()
    )

    row_idx = 2
    for r in rows_summary:
        wd = r.sale_date.weekday()
        days_es = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
        day_str = days_es[wd]
        
        is_payday = "Si" if r.sale_date.day in {14, 15, 28, 29, 30, 31} else "No"
        
        row_data = [
            r.sale_date,
            round(r.total_units, 1),
            round(r.total_revenue, 2),
            day_str,
            is_payday
        ]
        ws1.append(row_data)
        
        ws1.row_dimensions[row_idx].height = 19
        
        # Fecha
        cell_date = ws1.cell(row=row_idx, column=1)
        cell_date.number_format = 'yyyy-mm-dd'
        cell_date.alignment = align_center
        
        # Unidades
        cell_units = ws1.cell(row=row_idx, column=2)
        cell_units.number_format = '#,##0.0'
        cell_units.alignment = align_right
        
        # Ingresos
        cell_rev = ws1.cell(row=row_idx, column=3)
        cell_rev.number_format = '$#,##0.00'
        cell_rev.alignment = align_right
        
        # Dia y Quincena
        ws1.cell(row=row_idx, column=4).alignment = align_center
        ws1.cell(row=row_idx, column=5).alignment = align_center
        
        # Estilos generales
        for col_idx in range(1, len(headers1) + 1):
            c = ws1.cell(row=row_idx, column=col_idx)
            c.font = font_body
            c.border = border_data
            if row_idx % 2 == 1:
                c.fill = fill_zebra
                
        row_idx += 1

    # Fila de Totales Generales en Hoja 1
    if row_idx > 2:
        ws1.cell(row=row_idx, column=1, value="TOTAL GENERAL")
        ws1.cell(row=row_idx, column=2, value=f"=SUM(B2:B{row_idx-1})")
        ws1.cell(row=row_idx, column=3, value=f"=SUM(C2:C{row_idx-1})")
        
        ws1.row_dimensions[row_idx].height = 22
        
        ws1.cell(row=row_idx, column=1).alignment = align_left
        ws1.cell(row=row_idx, column=2).number_format = '#,##0.0'
        ws1.cell(row=row_idx, column=2).alignment = align_right
        ws1.cell(row=row_idx, column=3).number_format = '$#,##0.00'
        ws1.cell(row=row_idx, column=3).alignment = align_right
        
        for col_idx in range(1, len(headers1) + 1):
            c = ws1.cell(row=row_idx, column=col_idx)
            c.font = font_total
            c.fill = fill_total
            c.border = border_total
            
    # Autoajustar anchos ws1
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format == '$#,##0.00' and isinstance(cell.value, (int, float)):
                val = f"${cell.value:,.2f}"
            if len(val) > max_len:
                max_len = len(val)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)


    # -- Hoja 2: Detalle por Producto --
    ws2 = wb.create_sheet(title="Detalle por Producto")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Fecha", "Categoria", "Producto", "Unidades Vendidas", "Precio Unitario", "Ingresos Totales"]
    ws2.append(headers2)
    
    ws2.row_dimensions[1].height = 26
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Query detallado
    rows_detail = (
        db.query(
            SaleTransaction.sale_date,
            Product.category,
            Product.name.label("product_name"),
            SaleTransaction.quantity_sold,
            Product.selling_price,
            SaleTransaction.revenue
        )
        .join(Product, Product.id == SaleTransaction.product_id)
        .filter(SaleTransaction.scenario_id == scenario_id)
        .order_by(desc(SaleTransaction.sale_date), desc(SaleTransaction.quantity_sold))
        .all()
    )

    row_idx2 = 2
    for r in rows_detail:
        row_data2 = [
            r.sale_date,
            r.category,
            r.product_name,
            round(r.quantity_sold, 1),
            round(r.selling_price, 2),
            round(r.revenue, 2)
        ]
        ws2.append(row_data2)
        
        ws2.row_dimensions[row_idx2].height = 19
        
        # Formatos
        ws2.cell(row=row_idx2, column=1).number_format = 'yyyy-mm-dd'
        ws2.cell(row=row_idx2, column=1).alignment = align_center
        
        ws2.cell(row=row_idx2, column=2).alignment = align_left
        ws2.cell(row=row_idx2, column=3).alignment = align_left
        
        ws2.cell(row=row_idx2, column=4).number_format = '#,##0.0'
        ws2.cell(row=row_idx2, column=4).alignment = align_right
        
        ws2.cell(row=row_idx2, column=5).number_format = '$#,##0.00'
        ws2.cell(row=row_idx2, column=5).alignment = align_right
        
        ws2.cell(row=row_idx2, column=6).number_format = '$#,##0.00'
        ws2.cell(row=row_idx2, column=6).alignment = align_right
        
        for col_idx in range(1, len(headers2) + 1):
            c = ws2.cell(row=row_idx2, column=col_idx)
            c.font = font_body
            c.border = border_data
            if row_idx2 % 2 == 1:
                c.fill = fill_zebra
                
        row_idx2 += 1

    # Fila de Totales en Hoja 2
    if row_idx2 > 2:
        ws2.cell(row=row_idx2, column=1, value="TOTAL GENERAL")
        ws2.cell(row=row_idx2, column=4, value=f"=SUM(D2:D{row_idx2-1})")
        ws2.cell(row=row_idx2, column=6, value=f"=SUM(F2:F{row_idx2-1})")
        
        ws2.row_dimensions[row_idx2].height = 22
        
        ws2.cell(row=row_idx2, column=1).alignment = align_left
        ws2.cell(row=row_idx2, column=4).number_format = '#,##0.0'
        ws2.cell(row=row_idx2, column=4).alignment = align_right
        ws2.cell(row=row_idx2, column=6).number_format = '$#,##0.00'
        ws2.cell(row=row_idx2, column=6).alignment = align_right
        
        for col_idx in range(1, len(headers2) + 1):
            c = ws2.cell(row=row_idx2, column=col_idx)
            c.font = font_total
            c.fill = fill_total
            c.border = border_total

    # Autoajustar ws2
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format == '$#,##0.00' and isinstance(cell.value, (int, float)):
                val = f"${cell.value:,.2f}"
            if len(val) > max_len:
                max_len = len(val)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -- Retornar stream --
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
